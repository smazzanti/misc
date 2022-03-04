import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def format_value(value):
    if type(value) == bool:
        return str(value)
    try:
        float(value)
        return value
    except:
        return str(value)

def df_doc_2_excel(df, doc, filename, to_excel_args={"index": True}):

    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmpfile:
        df_format = df.applymap(format_value)
        df_format.to_excel(tmpfile.name, **to_excel_args)
        wb = load_workbook(tmpfile.name)

    ws_output = wb.active
    ws_output.title = 'Output'
    for col_ix in range(1, ws_output.max_column+1):
        ws_output.column_dimensions[get_column_letter(col_ix)].width = max(
            5, len(str(ws_output.cell(row=1, column=col_ix).value)))

    ws_doc = wb.create_sheet("Doc")
    ws_doc['B2'] = doc
    ws_doc['B2'].alignment = Alignment(wrapText=True)
    ws_doc.column_dimensions['B'].width = max([len(row) for row in doc.split('\n')]) * 1.25

    wb.save(filename)

def query_2_excel(query, bq_client, filename, to_excel_args={"index": False}):
    df = bq_client.query(query).to_dataframe()
    df_doc_2_excel(df=df, doc=query, filename=filename, to_excel_args=to_excel_args)
